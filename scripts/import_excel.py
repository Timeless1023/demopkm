#!/usr/bin/env python3
"""
Excel 导入脚本（MVP）

支持两个 sheet：
1) cards
   - card_name (必填)
   - card_code, set_name, rarity (可选)
2) purchases
   - card_name (必填)
   - qty (必填)
   - unit_cost (必填)
   - purchased_at (必填, e.g. 2026-04-21T10:00:00)
   - source (可选)
"""

from __future__ import annotations

import argparse
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.main import init_db


def get_db(db_path: Path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def parse_datetime(value) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        return datetime.fromisoformat(value).isoformat()
    raise ValueError(f"Invalid datetime value: {value}")


def import_cards(conn: sqlite3.Connection, ws) -> int:
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0
    for row in rows:
        card_name, card_code, set_name, rarity = row[:4]
        if not card_name:
            continue
        try:
            conn.execute(
                "INSERT INTO cards (card_name, card_code, set_name, rarity) VALUES (?, ?, ?, ?)",
                (str(card_name).strip(), card_code, set_name, rarity),
            )
            count += 1
        except sqlite3.IntegrityError:
            pass
    return count


def get_card_id_by_name(conn: sqlite3.Connection, card_name: str) -> int:
    row = conn.execute("SELECT id FROM cards WHERE card_name = ?", (card_name,)).fetchone()
    if not row:
        cur = conn.execute("INSERT INTO cards (card_name) VALUES (?)", (card_name,))
        return cur.lastrowid
    return row["id"]


def import_purchases(conn: sqlite3.Connection, ws) -> int:
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0
    for row in rows:
        card_name, qty, unit_cost, purchased_at, source = row[:5]
        if not card_name:
            continue
        card_id = get_card_id_by_name(conn, str(card_name).strip())
        qty_int = int(qty)
        unit_cost_float = float(unit_cost)
        purchased_at_iso = parse_datetime(purchased_at)
        conn.execute(
            """
            INSERT INTO inventory_lots (card_id, qty_in, qty_remaining, unit_cost, purchased_at, source)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (card_id, qty_int, qty_int, unit_cost_float, purchased_at_iso, source),
        )
        count += 1
    return count


def main():
    parser = argparse.ArgumentParser(description="Import cards and purchases from Excel")
    parser.add_argument("--file", required=True, help="Excel file path")
    parser.add_argument("--db", default="pokemon_trading.db", help="SQLite db path")
    args = parser.parse_args()

    excel_path = Path(args.file)
    db_path = Path(args.db)

    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    init_db(db_path)

    wb = load_workbook(excel_path)

    with get_db(db_path) as conn:
        cards_count = 0
        purchases_count = 0

        if "cards" in wb.sheetnames:
            cards_count = import_cards(conn, wb["cards"])

        if "purchases" in wb.sheetnames:
            purchases_count = import_purchases(conn, wb["purchases"])

        conn.commit()

    print(f"Imported cards: {cards_count}")
    print(f"Imported purchases: {purchases_count}")


if __name__ == "__main__":
    main()
